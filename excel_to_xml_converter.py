import os
from openpyxl import load_workbook
from lxml import etree


def convert_excel_to_xml(excel_file, xml_file):
    workbook = load_workbook(excel_file)

    root = etree.Element('workbook')

    for worksheet in workbook.worksheets:
        worksheet_element = etree.SubElement(
            root, 'worksheet', name=worksheet.title)

        for row in worksheet.iter_rows(values_only=True):
            row_element = etree.SubElement(worksheet_element, 'row')
            for cell in row:
                cell_element = etree.SubElement(row_element, 'cell')
                cell_element.text = str(cell)

    tree = etree.ElementTree(root)
    tree.write(xml_file, pretty_print=True,
               xml_declaration=True, encoding='UTF-8')


def main(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            excel_file = os.path.join(folder_path, filename)
            xml_file = os.path.splitext(excel_file)[0] + '.xml'
            convert_excel_to_xml(excel_file, xml_file)
            print(f"Converted {filename} to {os.path.basename(xml_file)}")
