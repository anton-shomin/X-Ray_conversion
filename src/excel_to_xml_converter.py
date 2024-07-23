import os
from openpyxl import load_workbook
from lxml import etree


def convert_excel_to_xml(excel_file, xml_file):
  """
  Convert an Excel file to an XML file.

  Args:
      excel_file (str): The path to the Excel file.
      xml_file (str): The path to the output XML file.

  This function reads an Excel file using the `load_workbook` function from the `openpyxl` library.
  It creates an XML root element with the tag 'workbook'.
  For each worksheet in the Excel file, it creates a subelement 'worksheet' with the worksheet title as an attribute.
  For each row in the worksheet, it creates a subelement 'row' and for each cell in the row, it creates a subelement 'cell' with the cell value as text.
  Finally, it writes the XML tree to the output XML file using the `write` method of the `ElementTree` class from the `lxml` library.

  Returns:
      None
  """
  workbook = load_workbook(excel_file)

  root = etree.Element('workbook')

  for worksheet in workbook.worksheets:
    worksheet_element = etree.SubElement(
      root, 'worksheet', name=worksheet.title)

    for row in worksheet.iter_rows(values_only=True):
      row_element = etree.SubElement(worksheet_element, 'row')
      for cell in row:
        cell_element = etree.SubElement(row_element, 'cell')
        cell_element.text = str(cell)

  tree = etree.ElementTree(root)
  tree.write(xml_file, pretty_print=True,
             xml_declaration=True, encoding='UTF-8')


def main(folder_path):
  for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
      excel_file = os.path.join(folder_path, filename)
      xml_file = os.path.splitext(excel_file)[0] + '.xml'
      convert_excel_to_xml(excel_file, xml_file)
      print(f"Converted {filename} to {os.path.basename(xml_file)}")
